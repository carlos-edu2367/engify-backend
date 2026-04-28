from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class CommissionReportRow:
    obra_id: UUID
    titulo: str
    valor_total: Decimal
    ultimo_recebimento: datetime


class ExcelReportBuilder:
    def build_monthly_commission_report(
        self,
        team_id: UUID,
        categoria_id: UUID,
        categoria_nome: str,
        mes: int,
        ano: int,
        porcentagem_comissao: Decimal,
        rows: list[CommissionReportRow],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio"

        self._build_header(
            sheet=sheet,
            categoria_nome=categoria_nome,
            mes=mes,
            ano=ano,
            porcentagem_comissao=porcentagem_comissao,
        )

        self._build_table(sheet, rows)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def load_workbook(self, content: bytes):
        return load_workbook(BytesIO(content))

    def _build_header(
        self,
        sheet,
        categoria_nome: str,
        mes: int,
        ano: int,
        porcentagem_comissao: Decimal,
    ) -> None:
        sheet["A1"] = "Relatorio financeiro de obras recebidas"
        sheet["A2"] = "Periodo"
        sheet["B2"] = f"{mes:02d}/{ano}"
        sheet["A3"] = "Comissao"
        sheet["B3"] = porcentagem_comissao
        sheet["D1"] = "Categoria"
        sheet["E1"] = categoria_nome

        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        accent_fill = PatternFill(fill_type="solid", fgColor="E5EEF9")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        for cell in ("A1",):
            sheet[cell].fill = header_fill
            sheet[cell].font = white_font
        for cell in ("A2", "A3", "D1"):
            sheet[cell].font = bold_font
            sheet[cell].fill = accent_fill

        sheet["B3"].number_format = "0.00%"

    def _build_table(self, sheet, rows: list[CommissionReportRow]) -> None:
        start_row = 5
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", fgColor="334155")
        total_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
        header_font = Font(color="FFFFFF", bold=True)
        total_font = Font(bold=True)

        sheet[f"A{start_row}"] = "Titulo da obra"
        sheet[f"B{start_row}"] = "Valor total da obra"
        sheet[f"C{start_row}"] = "Comissao"

        for cell_ref in (f"A{start_row}", f"B{start_row}", f"C{start_row}"):
            sheet[cell_ref].fill = header_fill
            sheet[cell_ref].font = header_font
            sheet[cell_ref].border = border

        data_row = start_row + 1
        for index, row in enumerate(rows, start=data_row):
            sheet[f"A{index}"] = _sanitize_excel_text(row.titulo)
            sheet[f"B{index}"] = row.valor_total
            sheet[f"C{index}"] = f"=B{index}*$B$3"

            sheet[f"A{index}"].border = border
            sheet[f"B{index}"].border = border
            sheet[f"C{index}"].border = border
            sheet[f"B{index}"].number_format = 'R$ #,##0.00'
            sheet[f"C{index}"].number_format = 'R$ #,##0.00'
            sheet[f"B{index}"].alignment = Alignment(horizontal="right")
            sheet[f"C{index}"].alignment = Alignment(horizontal="right")

        if rows:
            last_data_row = data_row + len(rows) - 1
            total_row = last_data_row + 2
            value_formula = f"=SUM(B{data_row}:B{last_data_row})"
            commission_formula = f"=SUM(C{data_row}:C{last_data_row})"
        else:
            sheet[f"A{data_row}"] = "Nenhuma obra elegivel no periodo"
            last_data_row = data_row
            total_row = data_row + 2
            value_formula = "0"
            commission_formula = "0"

        sheet[f"A{total_row}"] = "Total"
        sheet[f"B{total_row}"] = value_formula
        sheet[f"C{total_row}"] = commission_formula

        for cell_ref in (f"A{total_row}", f"B{total_row}", f"C{total_row}"):
            sheet[cell_ref].fill = total_fill
            sheet[cell_ref].font = total_font
            sheet[cell_ref].border = border

        sheet[f"B{total_row}"].number_format = 'R$ #,##0.00'
        sheet[f"C{total_row}"].number_format = 'R$ #,##0.00'
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:C{max(last_data_row, 6)}"
        sheet.column_dimensions["A"].width = 42
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 18
        sheet.column_dimensions["E"].width = 40


def _sanitize_excel_text(value: str) -> str:
    if value and value[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value
